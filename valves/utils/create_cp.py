import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill


def create_cp(cp_data):
  
  wb = Workbook()
  ws = wb.active

  # Data defining and locating =============================================================================================

  start_row_num = 7

  # Locates cp_data to table rows, adds formulas.
  for index, el in enumerate(cp_data):
    row_num = start_row_num + index

    ws[f'A{row_num}'] = index + 1
    ws[f'B{row_num}'] = el['code']
    ws[f'C{row_num}'] = f"{el['type_title']} {el['dn']}/{el['kvs']}" if el['equip_type'] in ['cv_valve', 'pr_valve',] else el['full_title']
    ws[f'D{row_num}'] = el['discount_group']
    ws[f'E{row_num}'] = el['amount']
    ws[f'F{row_num}'] = el['price']

    ws[f'G{row_num}'] = '=$C$3' if el['discount_group'] == '40RU PL08-DH-V' else '=$C$4'

    ws[f'H{row_num}'] = f'=ROUND(F{row_num}*((100-G{row_num})/100)*E{row_num}, 2)'
    ws[f'I{row_num}'] = f'=ROUND(H{row_num}*120%, 2)'

    print('=ROUND(F{row_num}*((100-G{row_num})/100)*E{row_num}, 2)')

  # Locates headers and formulas for total data in the last row.
  last_row_num = start_row_num + len(cp_data)

  ws[f'D{last_row_num}'] = 'Общее кол-во:'
  ws[f'E{last_row_num}'] = f'=SUBTOTAL(9, E{start_row_num}:E{last_row_num - 1})'
  
  ws[f'G{last_row_num}'] = 'Итого:'
  ws[f'H{last_row_num}'] = f'=SUBTOTAL(9, H{start_row_num}:H{last_row_num - 1})'
  ws[f'I{last_row_num}'] = f'=SUBTOTAL(9, I{start_row_num}:I{last_row_num - 1})'

  # ========================================================================================================================

  # Styling ================================================================================================================

  ws['B1'] = f'Пользовательская корзина, пересчитано: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'

  [ws['B2'], ws['C2'], ws['D2'],] = ['Группа скидок', 'Скидка клиента', 'МРЦ',]
  [ws['B3'], ws['C3'], ws['D3'],] = ['40RU PL08-DH-V', 0,               32,   ]
  [ws['B4'], ws['C4'], ws['D4'],] = ['28RU PL08-IWKB', 0,               23,   ]

  # Fill column headers.
  column_headers = ['п/н', 'Код', 'Описание', 'Группа скидок', 'Кол-во', 'Цена, EUR', 'Скидка', 'Сумма со скидкой, EUR', 'Итого с НДС, EUR',]
  for i, header in enumerate(column_headers):
    ws[f'{get_column_letter(i + 1)}6'] = header
  
  ws.merge_cells('B1:D1')

  # Sets columns' widths.
  columns_width = [3.5, 16, 35, 16, 7, 10.5, 7.5, 23, 17,]
  for i, column_width in enumerate(columns_width):
    ws.column_dimensions[get_column_letter(i + 1)].width = column_width

  # Fills cells' color.
  gray_cells = ws['B2:D2'] + ws['A6:I6']
  for a_tuple in gray_cells:
    for cell in a_tuple:
      cell.fill = PatternFill(start_color='999999', fill_type='solid')

  # Makes cells' text bold.
  bold_cells = ws['B2:D2'] + ws['A6:I6'] + ws[f'D{last_row_num}:I{last_row_num}']
  for a_tuple in bold_cells:
    for cell in a_tuple:
      cell.font = Font(bold=True)

  # ========================================================================================================================

  wb.save('files/cp.xlsx')
